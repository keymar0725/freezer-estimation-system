import pandas as pd
import openpyxl
import os

def load_input_data(filename="積算入力シート.xlsx"):
    """入力用Excelから基本情報を読み込む"""
    df = pd.read_excel(filename, sheet_name='基本情報・フリーザー仕様')
    return dict(zip(df['項目名'], df['入力欄']))

def fetch_unit_price(db_df, item_name, spec):
    """
    マニュアルDBから単価を検索して取得する関数
    ※実際の列名（名    称, 摘    要など）に合わせて検索します
    """
    try:
        # 名称と摘要（仕様）が部分一致する行を検索
        # ※実際のDBファイルの列名に合わせて '名    称' や '摘    要' を調整してください
        mask = db_df.iloc[:, 1].astype(str).str.contains(item_name, na=False) & \
               db_df.iloc[:, 2].astype(str).str.contains(spec, na=False)
        result = db_df[mask]
        
        if not result.empty:
            # 見つかった場合、複合単価（一番右の列など）を取得
            # 列インデックスは実際のファイルに合わせて調整します（例として10列目と仮定）
            price = result.iloc[0, 10] 
            # 文字列カンマ除去
            if isinstance(price, str):
                price = float(price.replace(',', ''))
            return price
        else:
            return 0 # 見つからない場合は0
    except Exception as e:
        print(f"単価検索エラー: {e}")
        return 0

def write_to_format(output_data, template_file="01-2 工事原価表.xlsx", project_name="新規案件"):
    """
    計算結果を工事原価表のフォーマットに書き込む
    """
    if not os.path.exists(template_file):
        print(f"エラー: テンプレートファイル '{template_file}' が見つかりません。")
        return

    wb = openpyxl.load_workbook(template_file)
    ws = wb.active

    # B列（列インデックス2）のカテゴリ名を探して、その下に書き込む簡易ロジック
    # output_data の形式: {"熱絶縁工事": [["パネル", "不燃100t", 150.5, "㎡", 10000, 1505000], ...]}
    
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=2).value
        
        if cell_val in output_data:
            items = output_data[cell_val]
            # カテゴリが見つかったら、その次の行からアイテムを書き込む
            for i, item in enumerate(items):
                write_row = row + 1 + i
                # B列: 項目, C列: 詳細, D列: 数量, E列: 単位, F列: 単価, G列: 金額
                ws.cell(row=write_row, column=2).value = item[0] # 項目
                ws.cell(row=write_row, column=3).value = item[1] # 詳細
                ws.cell(row=write_row, column=4).value = item[2] # 数量
                ws.cell(row=write_row, column=5).value = item[3] # 単位
                ws.cell(row=write_row, column=6).value = item[4] # 単価
                ws.cell(row=write_row, column=7).value = item[5] # 金額

    # 宛先や件名の書き込み（1行目、2行目付近）
    ws.cell(row=2, column=2).value = f"{project_name} 設置工事"

    output_filename = f"{project_name}_工事原価表.xlsx"
    wb.save(output_filename)
    print(f"\n🎉 成功: '{output_filename}' を出力しました！")

def calculate_and_export():
    # 1. 入力データの読み込み
    data = load_input_data()
    project_name = data.get('プロジェクト名', '名称未設定')
    print(f"=== 📊 積算実行: {project_name} ===")

    # 2. マニュアルDBの読み込み（初回のみ時間がかかります）
    db_file = "機械設備工事積算実務マニュアル2025_大阪府.xlsx"
    print("データベースを読み込んでいます...")
    try:
        # シート名が分からない場合は最初のシートを読み込む
        db_df = pd.read_excel(db_file, sheet_name=0, header=0)
    except Exception as e:
        print(f"DB読み込みエラー: {e}")
        db_df = None

    # 3. 数量の計算
    width = float(data['機巾 [mm]']) / 1000
    length = float(data['機長 [mm]']) / 1000
    height = float(data['機高 [mm]']) / 1000
    total_area = (width * length * 2) + (width * height * 2) + (length * height * 2)
    door_area = (float(data['扉の幅 [mm]']) / 1000) * (float(data['扉の高さ [mm]']) / 1000) * float(data['扉の枚数 [枚]'])
    net_panel_area = total_area - door_area

    # 4. 単価の取得と金額計算（DBがあれば検索、なければ仮単価）
    panel_unit_price = 0
    if db_df is not None:
        # 例：「防熱パネル」などのキーワードで検索（実際のDBの表記に合わせて調整します）
        # 今回は仮で固定値を入れるか、見つからなかった場合用のデフォルト値を設定
        panel_unit_price = fetch_unit_price(db_df, "パネル", "100t") 
    
    if panel_unit_price == 0:
        panel_unit_price = 15000 # DBから取得できなかった場合の仮単価

    panel_total_price = net_panel_area * panel_unit_price

    # 5. 原価表に書き込むデータ構造を作成
    output_data = {
        "フリーザー": [
            [f"フリーザー本体 ({data['フリーザー型式']})", "一式", 1, "式", 15000000, 15000000]
        ],
        "熱絶縁工事": [
            ["防熱パネル", f"厚さ(自動判定) 面積", round(net_panel_area, 1), "㎡", panel_unit_price, round(panel_total_price)]
        ]
    }

    # 水冷式配管の処理
    if str(data['凝縮方式']).strip() == "水冷式":
        straight_length = float(data['配管直管距離(往復) [m]'])
        multiplier = {"シンプル": 1.3, "普通": 1.5, "複雑": 2.0}.get(str(data['配管ルート複雑さ']).strip(), 1.5)
        eq_length = straight_length * multiplier
        
        # 配管単価の取得シミュレーション
        pipe_unit_price = fetch_unit_price(db_df, "塩ビライニング鋼管", "50A") if db_df is not None else 0
        if pipe_unit_price == 0: pipe_unit_price = 4500 # 仮単価

        output_data["設備工事"] = [
            ["冷却水配管", "50A (相当長換算)", round(eq_length, 1), "m", pipe_unit_price, round(eq_length * pipe_unit_price)],
            ["冷却塔", "能力自動選定", 1, "台", 800000, 800000]
        ]

    # 6. フォーマットへ出力
    write_to_format(output_data, project_name=project_name)

if __name__ == "__main__":
    calculate_and_export()
